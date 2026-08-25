import re
import csv
from dataclasses import dataclass
from io import StringIO
from pathlib import Path
from uuid import UUID

from openpyxl import load_workbook
from pypdf import PdfReader

from .upload import MAX_QUOTE_FILES, UPLOAD_DIRECTORY

SAVED_QUOTE_PATTERN = re.compile(
    r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-"
    r"[0-9a-fA-F]{4}-[0-9a-fA-F]{12}\.(pdf|xlsx|csv)$"
)


@dataclass(frozen=True)
class ExtractionFileRequest:
    original_name: str
    saved_name: str


@dataclass(frozen=True)
class ExtractionResult:
    original_name: str
    saved_name: str
    status: str
    page_count: int
    character_count: int
    extracted_text: str
    error_message: str | None


class ExtractionRequestError(Exception):
    def __init__(self, *, status_code: int, code: str, message: str) -> None:
        super().__init__(message)
        self.status_code = status_code
        self.code = code
        self.message = message


def _clean_page_text(text: str) -> str:
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    cleaned: list[str] = []
    previous_was_blank = False

    for line in lines:
        normalized = line.strip()
        is_blank = not normalized
        if is_blank and previous_was_blank:
            continue
        cleaned.append(normalized)
        previous_was_blank = is_blank

    return "\n".join(cleaned).strip()


def _resolve_uploaded_file(saved_name: str) -> Path:
    if (
        not saved_name
        or Path(saved_name).name != saved_name
        or "/" in saved_name
        or "\\" in saved_name
        or not SAVED_QUOTE_PATTERN.fullmatch(saved_name)
    ):
        raise ExtractionRequestError(
            status_code=400,
            code="INVALID_SAVED_NAME",
            message="저장 파일 이름 형식을 확인해 주세요.",
        )

    try:
        UUID(Path(saved_name).stem)
    except ValueError as error:
        raise ExtractionRequestError(
            status_code=400,
            code="INVALID_SAVED_NAME",
            message="저장 파일 이름 형식을 확인해 주세요.",
        ) from error

    upload_directory = UPLOAD_DIRECTORY.resolve()
    file_path = (upload_directory / saved_name).resolve()
    if file_path.parent != upload_directory:
        raise ExtractionRequestError(
            status_code=400,
            code="INVALID_FILE_PATH",
            message="허용되지 않은 파일 경로입니다.",
        )
    if not file_path.is_file():
        raise ExtractionRequestError(
            status_code=404,
            code="UPLOADED_FILE_NOT_FOUND",
            message="업로드된 견적서 파일을 찾을 수 없습니다.",
        )
    return file_path


def _failed_result(
    file: ExtractionFileRequest, message: str, page_count: int = 0
) -> ExtractionResult:
    return ExtractionResult(
        original_name=file.original_name,
        saved_name=file.saved_name,
        status="failed",
        page_count=page_count,
        character_count=0,
        extracted_text="",
        error_message=message,
    )


def extract_pdf_text(
    file: ExtractionFileRequest, file_path: Path
) -> ExtractionResult:
    page_count = 0
    try:
        reader = PdfReader(file_path)
        if reader.is_encrypted:
            return _failed_result(file, "암호화된 PDF는 텍스트를 추출할 수 없습니다.")

        page_count = len(reader.pages)
        pages: list[str] = []
        has_text = False
        for page_number, page in enumerate(reader.pages, start=1):
            page_text = _clean_page_text(page.extract_text() or "")
            if page_text:
                has_text = True
            pages.append(f"--- Page {page_number} ---\n{page_text}".rstrip())

        if not has_text:
            return ExtractionResult(
                original_name=file.original_name,
                saved_name=file.saved_name,
                status="needs_ocr",
                page_count=page_count,
                character_count=0,
                extracted_text="",
                error_message="스캔 문서로 추정되어 텍스트를 읽을 수 없습니다.",
            )

        extracted_text = "\n\n".join(pages).strip()
        return ExtractionResult(
            original_name=file.original_name,
            saved_name=file.saved_name,
            status="success",
            page_count=page_count,
            character_count=len(extracted_text),
            extracted_text=extracted_text,
            error_message=None,
        )
    except Exception:
        return _failed_result(
            file,
            "PDF 파일을 열거나 텍스트를 추출하는 중 오류가 발생했습니다.",
            page_count,
        )


def _table_text(rows: list[list[object]]) -> str:
    lines = []
    for row in rows:
        values = [str(value).strip() if value is not None else "" for value in row]
        if any(values):
            lines.append("\t".join(values).rstrip())
    return "\n".join(lines).strip()


def extract_csv_text(file: ExtractionFileRequest, file_path: Path) -> ExtractionResult:
    try:
        raw = file_path.read_bytes()
        decoded = None
        for encoding in ("utf-8-sig", "cp949"):
            try:
                decoded = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if decoded is None:
            return _failed_result(file, "CSV 문자 인코딩을 확인할 수 없습니다.")
        rows = [list(row) for row in csv.reader(StringIO(decoded))]
        extracted_text = _table_text(rows)
        if not extracted_text:
            return _failed_result(file, "CSV에서 확인할 수 있는 데이터가 없습니다.", 1)
        return ExtractionResult(
            original_name=file.original_name,
            saved_name=file.saved_name,
            status="success",
            page_count=1,
            character_count=len(extracted_text),
            extracted_text=extracted_text,
            error_message=None,
        )
    except Exception:
        return _failed_result(file, "CSV 파일을 읽는 중 오류가 발생했습니다.")


def extract_xlsx_text(file: ExtractionFileRequest, file_path: Path) -> ExtractionResult:
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheets: list[str] = []
        for worksheet in workbook.worksheets:
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            sheet_text = _table_text(rows)
            sheets.append(f"--- Sheet: {worksheet.title} ---\n{sheet_text}".rstrip())
        workbook.close()
        extracted_text = "\n\n".join(sheets).strip()
        if not any(section.partition("\n")[2].strip() for section in sheets):
            return _failed_result(file, "XLSX에서 확인할 수 있는 데이터가 없습니다.", len(sheets))
        return ExtractionResult(
            original_name=file.original_name,
            saved_name=file.saved_name,
            status="success",
            page_count=len(sheets),
            character_count=len(extracted_text),
            extracted_text=extracted_text,
            error_message=None,
        )
    except Exception:
        return _failed_result(file, "XLSX 파일을 읽는 중 오류가 발생했습니다.")


def extract_uploaded_pdfs(
    files: list[ExtractionFileRequest],
) -> list[ExtractionResult]:
    if not files:
        raise ExtractionRequestError(
            status_code=400,
            code="FILES_REQUIRED",
            message="텍스트를 추출할 견적서 파일을 한 개 이상 선택해 주세요.",
        )
    if len(files) > MAX_QUOTE_FILES:
        raise ExtractionRequestError(
            status_code=400,
            code="TOO_MANY_FILES",
            message="견적서 텍스트 추출은 최대 5개까지 요청할 수 있습니다.",
        )

    resolved = [(file, _resolve_uploaded_file(file.saved_name)) for file in files]
    extractors = {
        ".pdf": extract_pdf_text,
        ".csv": extract_csv_text,
        ".xlsx": extract_xlsx_text,
    }
    return [extractors[file_path.suffix.lower()](file, file_path) for file, file_path in resolved]
